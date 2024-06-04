from enum import Enum

from openpyxl import load_workbook

from countries_mapper import (
    CountryNameSolver,
    territories_regions_unrecognized_countries,
)
from db import Session
from models.country_statistics import CountryStatistics
from models.education_dimension import EducationDimension
from models.time_dimension import TimeDimension

DATA_RANGE = "B13:D796"


class ReportType(Enum):
    MATH = 1
    SCIENCE = 2
    READING = 3


REPORT_TYPE_MAP = {
    ReportType.MATH: "math_score",
    ReportType.SCIENCE: "science_score",
    ReportType.READING: "reading_score",
}


class EducationDimensionLoadError(Exception):
    pass


def load_education_dimension():
    excel = load_workbook("etl/data/pisa_scores.xlsx")

    # Data has been divided into three worksheets: Math, Science, and Reading
    # So we need to link data by year and country
    education_dimension_dict = {}
    for worksheet_name, report_type in zip(excel.sheetnames, list(ReportType)):
        worksheet = excel[worksheet_name]

        year = None
        for row in worksheet[DATA_RANGE]:
            if row[0].value is not None and row[0].value.isdecimal():
                year = int(row[0].value)

            country_name = row[1].value
            if country_name.lower() in territories_regions_unrecognized_countries:
                continue

            dict_key = (year, country_name)
            if dict_key not in education_dimension_dict:
                try:
                    education_dimension_dict[dict_key] = {REPORT_TYPE_MAP[report_type]: int(row[2].value)}
                except ValueError:
                    education_dimension_dict[dict_key] = {REPORT_TYPE_MAP[report_type]: 0}
            else:
                try:
                    education_dimension_dict[dict_key][REPORT_TYPE_MAP[report_type]] = int(row[2].value)
                except ValueError:
                    education_dimension_dict[dict_key][REPORT_TYPE_MAP[report_type]] = 0

    # create needed time dimension objects
    data_years = set([year for year, _ in education_dimension_dict])
    with Session() as session:
        existed_years = set(session.query(TimeDimension.year).all())
        not_existed_years = data_years - existed_years

        to_create_time_dimension = [TimeDimension(year=year) for year in not_existed_years]
        session.add_all(to_create_time_dimension)

        session.commit()

    with Session() as session:
        year_2_time_id = dict(session.query(TimeDimension.year, TimeDimension.time_id).all())

    # parse data and load it to the database + link it to the country statistics
    for year, country in education_dimension_dict:
        country_name_solver = CountryNameSolver()
        country_id = country_name_solver.solve(country)
        try:
            if country_id is None:
                raise EducationDimensionLoadError(f"Country {country} not found in the database")
        except EducationDimensionLoadError as e:
            print(e)
            continue

        time_id = year_2_time_id[year]

        with Session() as session:
            country_statistics = (
                session.query(CountryStatistics)
                .filter(CountryStatistics.country_id == country_id, CountryStatistics.time_id == time_id)
                .one_or_none()
            )
            if country_statistics is None:
                country_statistics = CountryStatistics(
                    country_id=country_id,
                    time_id=time_id,
                )
                session.add(country_statistics)
            session.commit()
        education_dimension = EducationDimension(
            math_score=education_dimension_dict[(year, country)].get("math_score"),
            science_score=education_dimension_dict[(year, country)].get("science_score"),
            reading_score=education_dimension_dict[(year, country)].get("reading_score"),
            pisa_average_score=(
                education_dimension_dict[(year, country)].get("math_score")
                + education_dimension_dict[(year, country)].get("science_score")
                + education_dimension_dict[(year, country)].get("reading_score")
            )
            / 3,
        )

        with Session() as session:
            session.add(education_dimension)
            session.flush()
            session.refresh(education_dimension)
            country_statistics.education_id = education_dimension.education_id
            session.add(country_statistics)
            session.commit()
